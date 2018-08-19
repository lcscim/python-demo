import openpyxl
import os
from openpyxl.styles import Font
from tkinter import *
import tkinter.filedialog
import warnings
# warnings谨慎使用
warnings.filterwarnings("ignore")

root = tkinter.Tk()
default_dir = r"C:\Users\asus\Desktop"
fname = tkinter.filedialog.askopenfilename(title=u"选择文件",initialdir=(os.path.expanduser(default_dir)))

wb = openpyxl.load_workbook('{}'.format(fname))
root.withdraw()
ws = wb['订单列表']
rows = ws.max_row
names = []
for each in ws['C2:C{}'.format(rows-1)]:
    for each_cell in each:
        if each_cell.value not in names:
            names.append(each_cell.value)

def geshi(name,i):
    wb.create_sheet(index = i+1,title = name)
    wb[name].guess_types = True
    wb[name]['A1'] = '商品名称'
    wb[name]['B1'] = '商品数量'
    wb[name]['C1'] = '收货人'
    wb[name]['D1'] = '收货电话'
    wb[name]['E1'] = '收货地址'
    wb[name]['A1'].font = Font(bold=True, size=15)
    wb[name]['B1'].font = Font(bold=True, size=15)
    wb[name]['C1'].font = Font(bold=True, size=15)
    wb[name]['D1'].font = Font(bold=True, size=15)
    wb[name]['E1'].font = Font(bold=True, size=15)
    wb[name].column_dimensions['A'].width = 40
    wb[name].column_dimensions['E'].width = 50
    wb[name].column_dimensions['D'].width = 15
    j = 2
    for a in range(rows-2):
        
        if names[i] == ws['C{}'.format(a+2)].value:
            
            wb[name]['A{}'.format(j)] = ws['C{}'.format(a+2)].value
            wb[name]['B{}'.format(j)] = ws['E{}'.format(a+2)].value
            if ws['K{}'.format(a+2)].value != '':
                
                wb[name]['C{}'.format(j)] = ws['K{}'.format(a+2)].value
                wb[name]['D{}'.format(j)] = ws['L{}'.format(a+2)].value
                wb[name]['E{}'.format(j)] = ws['M{}'.format(a+2)].value
            else:
                k = 1
                while True:
                    if ws['K{}'.format(a+2-k)].value != '':
                        wb[name]['C{}'.format(j)] = ws['K{}'.format(a+2-k)].value
                        wb[name]['D{}'.format(j)] = ws['L{}'.format(a+2-k)].value
                        wb[name]['E{}'.format(j)] = ws['M{}'.format(a+2-k)].value
                        break
                    else:
                        k += 1
            j+=1
        else:
            continue
    wb.save('{}'.format(fname))

for i in range(len(names)):
    if len(names[i]) < 30:
        name = names[i]
        if len(wb.sheetnames) <= len(names)+1:
            geshi(name,i)
        else:
            print('已录入')
    else:
        name = names[i][:20]+' '+names[i][30:]
        if len(wb.sheetnames) <= len(names)+1:
            geshi(name,i)
        else:
            print('已录入')
    
'''
a = [{'供应商':'张三','商品':['黑贝南瓜 39.9元5斤包邮','红蜜薯5斤39元包邮 48小时内发货']},
     {'供应商':'张三','商品':['黑贝南瓜 39.9元5斤包邮','红蜜薯5斤39元包邮 48小时内发货']}]

'''
            


