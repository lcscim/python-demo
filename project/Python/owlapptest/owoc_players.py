#__author__:asus
#date:2018/12/22
import requests
import openpyxl
from openpyxl.styles import Font,Alignment,NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string


def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchcontenders.com/teams"
    teams = requests.get(url=heros_url, headers=headers).json().get('competitors')
    teams_num = len(teams)
    print('一共有%d个队员' % teams_num)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(highlight)
    wb.create_sheet(index=0,title="players")
    ws = wb["players"]
    ws['A1'] = "序号"
    ws['B1'] = "队伍id"
    ws['C1'] = "队伍名称"
    ws['D1'] = "队伍简称"
    ws['E1'] = "赛区"
    ws['F1'] = "队员id"
    ws['G1'] = "队员名字"
    ws['H1'] = "队员姓名"
    ws['I1'] = "队员国籍"
    ws['J1'] = "primaryColor"
    ws['K1'] = "secondaryColor"
    ws.auto_filter.ref = "A1:K1"
        
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['D1'].style = highlight
    ws['E1'].style = highlight
    ws['F1'].style = highlight
    ws['G1'].style = highlight
    ws['H1'].style = highlight
    ws['I1'].style = highlight
    ws['J1'].style = highlight
    ws['K1'].style = highlight
        
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 6
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 9
        
    i = 2
    for team in teams:

        for player in team.get("competitor").get("players"):
            team_id = team.get("competitor").get("id")
            team_name = team.get("competitor").get("name")
            abbreviatedName = team.get("competitor").get("abbreviatedName")
            region = team.get("competitor").get("region")
            player_id = player.get('player').get("id")
            player_name = player.get('player').get("name")
            name = "{} {}".format(player.get('player').get("givenName"),player.get('player').get("familyName"))
            nationality = player.get('player').get("nationality")
            primaryColor = "#{}".format(team.get("competitor").get("primaryColor"))
            secondaryColor = "#{}".format(team.get("competitor").get("secondaryColor"))
            ws['A{0}'.format(i)] = i-1
            ws['B{0}'.format(i)] = team_id
            ws['C{0}'.format(i)] = team_name
            ws['D{0}'.format(i)] = abbreviatedName
            ws['E{0}'.format(i)] = region
            ws['F{0}'.format(i)] = player_id
            ws['G{0}'.format(i)] = player_name
            ws['H{0}'.format(i)] = name
            ws['I{0}'.format(i)] = nationality
            ws['J{0}'.format(i)] = primaryColor
            ws['K{0}'.format(i)] = secondaryColor
            
            ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['G{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['H{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['I{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['J{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['K{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            
            i += 1
    
    wb.save("owoc_players.xlsx")

if __name__ == '__main__':
    save_excel()

