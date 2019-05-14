#__author__:asus
#date:2018/12/22
import requests
import openpyxl
import datetime
from openpyxl.styles import Font,Alignment,NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string

now = datetime.datetime.now().strftime('%Y-%m-%d')
sheets = ['综合数据','dva','orisa','reinhardt','roadhog','winston','zarya','wreckingball','ashe','bastion',
        'doomfist','genji','hanzo','junkrat','mccree','mei','pharah','reaper','soldier-76','sombra',
        'symmetra','torbjorn','tracer','widowmaker','ana','brigitte','lucio','mercy','moira','zenyatta','baptiste']

def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/players"
    players = requests.get(url=heros_url, headers=headers).json().get('content')
    player_num = len(players)
    print('一共有%d个队员' % player_num)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(highlight)
    
    for a in range(len(sheets)):
        wb.create_sheet(index = int("{}".format(a)),title="{}".format(sheets[a]))
        ws = wb["{}".format(sheets[a])]

        ws['A1'] = "序号"
        ws['B1'] = "player_id"
        ws['C1'] = "player_name"
        ws['D1'] = "role"
        ws['E1'] = "team_id"
        ws['F1'] = "team_name"
        ws['G1'] = "{}平均每10分钟消灭".format(sheets[a])
        ws['H1'] = "排名"
        ws['I1'] = "{}平均每10分钟死亡".format(sheets[a])
        ws['J1'] = "排名"
        ws['K1'] = "{}平均每10分钟伤害".format(sheets[a])
        ws['L1'] = "排名"
        ws['M1'] = "{}平均每10分钟治疗".format(sheets[a])
        ws['N1'] = "排名"
        ws['O1'] = "{}平均每10分钟最终收益".format(sheets[a])
        ws['P1'] = "排名"
        ws['Q1'] = "{}平均每10分钟最后一击".format(sheets[a])
        ws['R1'] = "排名"
        if a > 0:
            ws['S1'] = "{}游戏总时长(s)".format(sheets[a])
            ws.auto_filter.ref = "A1:S{}".format(player_num+1)
        else:
            ws['S1'] = "{}游戏总时长(min)".format(sheets[a])
            ws['T1'] = "使用英雄数量"
            ws['U1'] = "所用英雄（包括未打出伤害或使用时间短）"
            ws.merge_cells('U1:V1')
            ws['T1'].style = highlight
            ws['U1'].style = highlight
            ws.column_dimensions['T'].width = 6
            ws.auto_filter.ref = "A1:AX{}".format(player_num+1)
            
        ws['A1'].style = highlight
        ws['B1'].style = highlight
        ws['C1'].style = highlight
        ws['D1'].style = highlight
        ws['E1'].style = highlight
        ws['F1'].style = highlight
        ws['G1'].style = highlight
        ws['H1'].style = highlight
        ws['I1'].style = highlight
        ws['J1'].style = highlight
        ws['K1'].style = highlight
        ws['L1'].style = highlight
        ws['M1'].style = highlight
        ws['N1'].style = highlight
        ws['O1'].style = highlight
        ws['P1'].style = highlight
        ws['Q1'].style = highlight
        ws['R1'].style = highlight
        ws['S1'].style = highlight
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 6
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 6
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 6
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 6
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 6
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 6
        ws.column_dimensions['S'].width = 12
        
        i = 2
        
        for player in players:
            player_id = player.get('id')
            player_name = player.get('name')
            role = player.get('attributes').get('role')
            team_id = player.get('teams')[0].get('team').get('id')
            team_name = player.get('teams')[0].get('team').get('name')
            ws['A{0}'.format(i)] = i-1
            ws['B{0}'.format(i)] = player_id
            ws['C{0}'.format(i)] = player_name
            ws['D{0}'.format(i)] = role
            ws['E{0}'.format(i)] = team_id
            ws['F{0}'.format(i)] = team_name
            
            ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['G{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['H{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['I{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['J{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['K{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['L{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['M{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['N{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['O{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['P{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['Q{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['R{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['S{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            i += 1
    
    wb.save("players_data({}).xlsx".format(now))
def data(id):
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/players/{}?locale=zh-cn&season=2019&stage_id=regular_season&expand=stats,stat.ranks".format(id)
    datas = requests.get(url=heros_url, headers=headers).json().get('data')
    num = []
    
    player_name = datas.get('player').get('id')
    eliminations_avg_per_10m = datas.get('stats').get('all').get('eliminations_avg_per_10m')
    deaths_avg_per_10m = datas.get('stats').get('all').get('deaths_avg_per_10m')
    hero_damage_avg_per_10m = datas.get('stats').get('all').get('hero_damage_avg_per_10m')
    healing_avg_per_10m = datas.get('stats').get('all').get('healing_avg_per_10m')
    ultimates_earned_avg_per_10m = datas.get('stats').get('all').get('ultimates_earned_avg_per_10m')
    final_blows_avg_per_10m = datas.get('stats').get('all').get('final_blows_avg_per_10m')
    time_played_total = datas.get('stats').get('all').get('time_played_total')
    rank1 = datas.get('statRanks').get('eliminations_avg_per_10m')
    rank2 = datas.get('statRanks').get('deaths_avg_per_10m')
    rank3 = datas.get('statRanks').get('hero_damage_avg_per_10m')
    rank4 = datas.get('statRanks').get('healing_avg_per_10m')
    rank5 = datas.get('statRanks').get('ultimates_earned_avg_per_10m')
    rank6 = datas.get('statRanks').get('final_blows_avg_per_10m')
    heroes = datas.get('stats').get('heroes')
    
    hero_name = []
    hero_eliminations = []
    hero_deaths = []
    hero_damage = []
    hero_healing = []
    hero_ultimates_earned = []
    hero_final_blows = []
    hero_time_played_total = []
    for hero in heroes:
        hero_name.append(hero.get('name'))
        hero_eliminations.append(hero.get('stats').get('eliminations_avg_per_10m'))
        hero_deaths.append(hero.get('stats').get('deaths_avg_per_10m'))
        hero_damage.append(hero.get('stats').get('hero_damage_avg_per_10m'))
        hero_healing.append(hero.get('stats').get('healing_avg_per_10m'))
        hero_ultimates_earned.append(hero.get('stats').get('ultimates_earned_avg_per_10m'))
        hero_final_blows.append(hero.get('stats').get('final_blows_avg_per_10m'))
        hero_time_played_total.append(hero.get('stats').get('time_played_total'))
    num.append([player_name,[
        eliminations_avg_per_10m,
        deaths_avg_per_10m,
        hero_damage_avg_per_10m,
        healing_avg_per_10m,
        ultimates_earned_avg_per_10m,
        final_blows_avg_per_10m,
        time_played_total
    ],[rank1,rank2,rank3,rank4,rank5,rank6],
    hero_name,hero_eliminations,hero_deaths,hero_damage,
    hero_healing,hero_ultimates_earned,hero_final_blows,hero_time_played_total])
    return num

def other_msg():

    wb = openpyxl.load_workbook("players_data({}).xlsx".format(now))
    wb.guess_types = True
    rows = wb["综合数据"].max_row
    player_id = []
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/stats/players"
    players = requests.get(url=heros_url, headers=headers).json().get('data')
    for player in players:
        id = player.get('playerId')
        player_id.append(id)
    print('上过场的有%d个队员' % len(players))
    for i in range(rows):
        if i+1<2 & i<rows-1:
            continue
        else:
            for j in player_id:
                if wb["综合数据"].cell(row=int("{}".format(i+1)),column=2).value == j:
                    player_data = data(j)
                    print(j)
                    wb["综合数据"]['G{}'.format(i+1)] = '{}'.format(player_data[0][1][0])
                    wb["综合数据"]['H{}'.format(i+1)] = '{}'.format(player_data[0][2][0])
                    wb["综合数据"]['I{}'.format(i+1)] = '{}'.format(player_data[0][1][1])
                    wb["综合数据"]['J{}'.format(i+1)] = '{}'.format(player_data[0][2][1])
                    wb["综合数据"]['K{}'.format(i+1)] = '{}'.format(player_data[0][1][2])
                    wb["综合数据"]['L{}'.format(i+1)] = '{}'.format(player_data[0][2][2])
                    wb["综合数据"]['M{}'.format(i+1)] = '{}'.format(player_data[0][1][3])
                    wb["综合数据"]['N{}'.format(i+1)] = '{}'.format(player_data[0][2][3])
                    wb["综合数据"]['O{}'.format(i+1)] = '{}'.format(player_data[0][1][4])
                    wb["综合数据"]['P{}'.format(i+1)] = '{}'.format(player_data[0][2][4])
                    wb["综合数据"]['Q{}'.format(i+1)] = '{}'.format(player_data[0][1][5])
                    wb["综合数据"]['R{}'.format(i+1)] = '{}'.format(player_data[0][2][5])
                    wb["综合数据"]['S{}'.format(i+1)] = '{}'.format(player_data[0][1][6]/60)
                    wb["综合数据"]['T{}'.format(i+1)] = '{}'.format(len(player_data[0][3]))
                    wb["综合数据"]['T{}'.format(i+1)].alignment = Alignment(horizontal='center', vertical='center')
                    
                    cols = 21

                    for hero in player_data[0][3]:
                        
                        wb["综合数据"].cell(row=int("{}".format(i+1)),column=int("{}".format(cols)),value=hero)
                        wb["综合数据"]['{0}{1}'.format(get_column_letter(cols),i+1)].alignment = \
                           Alignment(horizontal='center', vertical='center')
                        hero_index = player_data[0][3].index(hero)
                        wb["{}".format(hero)]['G{}'.format(i+1)] = '{}'.format(player_data[0][4][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['I{}'.format(i+1)] = '{}'.format(player_data[0][5][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['K{}'.format(i+1)] = '{}'.format(player_data[0][6][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['M{}'.format(i+1)] = '{}'.format(player_data[0][7][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['O{}'.format(i+1)] = '{}'.format(player_data[0][8][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['Q{}'.format(i+1)] = '{}'.format(player_data[0][9][int("{}".format(hero_index))])
                        wb["{}".format(hero)]['S{}'.format(i+1)] = '{}'.format(player_data[0][10][int("{}".format(hero_index))])
                        cols+=1
                    break
                else:
                    continue
    wb.save("players_data({}).xlsx".format(now))
            

if __name__ == '__main__':
    save_excel()
    other_msg()

