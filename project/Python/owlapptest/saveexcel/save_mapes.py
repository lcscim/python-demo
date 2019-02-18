#__author__:asus
#date:2018/12/22
import requests
import openpyxl
from openpyxl.styles import Font,Alignment,NamedStyle

def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/maps"
    maps = requests.get(url=heros_url, headers=headers).json()
    maps_num = len(maps)
    print('一共有%d个地图' % maps_num)
    wb.create_sheet(index=0,title="maps")
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(highlight)
    wb.guess_types=True
    ws = wb["maps"]
    ws['A1'] = "序号"
    ws['B1'] = "guid"
    ws['C1'] = "en_US"
    ws['D1'] = "ko_KR"
    ws['E1'] = "zh_CN"
    ws['F1'] = "map_type"
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['D1'].style = highlight
    ws['E1'].style = highlight
    ws['F1'].style = highlight
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    i = 2
    for map in maps:
        guid = map.get('guid')
        en_US = map.get('name').get('en_US')
        ko_KR = map.get('name').get('ko_KR')
        zh_CN = map.get('name').get('zh_CN')
        map_type = map.get('type')
        ws['A{0}'.format(i)] = i - 1
        ws['B{0}'.format(i)] = guid
        ws['C{0}'.format(i)] = en_US
        ws['D{0}'.format(i)] = ko_KR
        ws['E{0}'.format(i)] = zh_CN
        ws['F{0}'.format(i)] = map_type
        ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        i+=1

    wb.save("maps.xlsx")


if __name__ == '__main__':
    save_excel()

