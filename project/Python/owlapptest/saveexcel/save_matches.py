#__author__:asus
#date:2018/12/22
import requests
import openpyxl
from openpyxl.styles import Font,Alignment,NamedStyle

def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/schedule?locale=zh-CN"
    stages = requests.get(url=heros_url, headers=headers).json().get('data').get('stages')
    stages_num = len(stages)
    print('一共有%d个阶段' % stages_num)
    wb.create_sheet(index=0,title="stages")
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(highlight)
    wb.guess_types=True
    ws = wb["stages"]
    ws['A1'] = "序号"
    ws['B1'] = "stage_name"
    ws['C1'] = "week_name"
    ws['D1'] = "team_a"
    ws['E1'] = "team_b"
    ws['F1'] = "match_1"
    ws['G1'] = "match_2"
    ws['H1'] = "match_3"
    ws['I1'] = "match_4"
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['D1'].style = highlight
    ws['E1'].style = highlight
    ws['F1'].style = highlight
    ws['G1'].style = highlight
    ws['H1'].style = highlight
    ws['I1'].style = highlight
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    i = 2
    for stage in stages:
        stage_name = stage.get('name')
        weeks = stage.get('weeks')
        for week in weeks:
            week_name = week.get('name')
            matches = week.get('matches')
            for match in matches:
                try:
                    team_a = match.get('competitors')[0].get('name')
                    team_b = match.get('competitors')[1].get('name')
                    match_1 = match.get('games')[0].get('attributes').get('mapGuid')
                    match_2 = match.get('games')[1].get('attributes').get('mapGuid')
                    match_3 = match.get('games')[2].get('attributes').get('mapGuid')
                    match_4 = match.get('games')[3].get('attributes').get('mapGuid')
                    ws['A{0}'.format(i)] = i - 1
                    ws['B{0}'.format(i)] = stage_name
                    ws['C{0}'.format(i)] = week_name
                    ws['D{0}'.format(i)] = team_a
                    ws['E{0}'.format(i)] = team_b
                    ws['F{0}'.format(i)] = match_1
                    ws['G{0}'.format(i)] = match_2
                    ws['H{0}'.format(i)] = match_3
                    ws['I{0}'.format(i)] = match_4
                    ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['G{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['H{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['I{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
                except IndexError:
                    continue
                i+=1

    wb.save("matches.xlsx")


if __name__ == '__main__':
    save_excel()

