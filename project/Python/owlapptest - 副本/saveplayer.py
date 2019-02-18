#__author__:asus
#date:2018/12/22
import requests
import openpyxl
from openpyxl.styles import Font,Alignment,NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string

def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/players"
    players = requests.get(url=heros_url, headers=headers).json().get('content')
    player_num = len(players)
    print('一共有%d个队员' % player_num)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(highlight)
    wb.create_sheet(index = 0,title="players")
    ws = wb["players"]
    ws['A1'] = "序号"
    ws['B1'] = "player_id"
    ws['C1'] = "player_name"
    ws['D1'] = "role"
    ws['E1'] = "team_id"
    ws['F1'] = "team_name"
    ws['G1'] = "平均每10分钟消灭"
    ws['H1'] = "平均每10分钟死亡"
    ws['I1'] = "平均每10分钟伤害"
    ws['J1'] = "平均每10分钟治疗"
    ws['K1'] = "平均每10分钟最终收益"
    ws['L1'] = "平均每10分钟最后一击"
    ws['M1'] = "平均每10分钟游戏总时长(min)"
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['D1'].style = highlight
    ws['E1'].style = highlight
    ws['F1'].style = highlight
    ws['G1'].style = highlight
    ws['H1'].style = highlight
    ws['I1'].style = highlight
    ws['J1'].style = highlight
    ws['K1'].style = highlight
    ws['L1'].style = highlight
    ws['M1'].style = highlight
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['J'].width = 20
    i = 2
    players_num = 0
    for player in players:
        player_id = player.get('id')
        player_name = player.get('name')
        role = player.get('attributes').get('role')
        team_id = player.get('teams')[0].get('team').get('id')
        team_name = player.get('teams')[0].get('team').get('name')
        ws['A{0}'.format(i)] = i-1
        ws['B{0}'.format(i)] = player_id
        ws['C{0}'.format(i)] = player_name
        ws['D{0}'.format(i)] = role
        ws['E{0}'.format(i)] = team_id
        ws['F{0}'.format(i)] = team_name
        
        ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['G{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['H{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['I{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['J{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['K{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['L{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws['M{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
        i += 1
    print('一共有%d个队员' % players_num)
    wb.save("saveplayer.xlsx")
def data():
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/stats/players"
    players = requests.get(url=heros_url, headers=headers).json().get('data')
    num = []
    for player in players:
        player_name = player.get('name')
        eliminations_avg_per_10m = float(player.get('eliminations_avg_per_10m'))
        deaths_avg_per_10m = float(player.get('deaths_avg_per_10m'))
        hero_damage_avg_per_10m = float(player.get('hero_damage_avg_per_10m'))
        healing_avg_per_10m = float(player.get('healing_avg_per_10m'))
        ultimates_earned_avg_per_10m = float(player.get('ultimates_earned_avg_per_10m'))
        final_blows_avg_per_10m = float(player.get('final_blows_avg_per_10m'))
        time_played_total = float(player.get('time_played_total'))
        num.append([player_name,[
            eliminations_avg_per_10m,
            deaths_avg_per_10m,
            hero_damage_avg_per_10m,
            healing_avg_per_10m,
            ultimates_earned_avg_per_10m,
            final_blows_avg_per_10m,
            time_played_total
        ]])
    return num

def other_msg():
    wb = openpyxl.load_workbook("saveplayer.xlsx")
    wb.guess_types = True
    ws = wb["players"]
    rows = ws.max_row
    for i in range(rows):
        if i + 1 < 2:
            continue
        else:
            for j in data():
                if ws.cell(row="{}".format(i+1),column='C').value == j[1]:
                    
            


if __name__ == '__main__':
    save_excel()

