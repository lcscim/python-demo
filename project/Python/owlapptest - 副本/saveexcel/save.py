#__author__:asus
#date:2018/12/22
import requests
import openpyxl
from openpyxl.styles import Font,Alignment,NamedStyle

def save_excel():
    wb = openpyxl.Workbook()
    wb.guess_types = True
    headers = {'Host': 'api.overwatchleague.cn',
               'Connection': 'Keep-Alive',
               'Accept-Encoding': 'gzip',
               'User-Agent': 'okhttp/3.6.0'
               }
    heros_url = "https://api.overwatchleague.cn/teams?expand=team.content&locale=zh_CN"
    teams = requests.get(url=heros_url, headers=headers).json().get('competitors')
    teams_num = len(teams)
    print('一共有%d个队伍' % teams_num)
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    highlight.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(highlight)
    wb.create_sheet(index = 0,title="players")
    ws = wb["players"]
    ws['A1'] = "序号"
    ws['B1'] = "player_id"
    ws['C1'] = "player_name"
    ws['D1'] = "homeLocation"
    ws['E1'] = "player_number"
    ws['F1'] = "player_role"
    ws['G1'] = "name"
    ws['H1'] = "nationality"
    ws['I1'] = "team_id"
    ws['J1'] = "team_name"
    ws['K1'] = "team_primaryColor"
    ws['L1'] = "team_secondaryColor"
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['D1'].style = highlight
    ws['E1'].style = highlight
    ws['F1'].style = highlight
    ws['G1'].style = highlight
    ws['H1'].style = highlight
    ws['I1'].style = highlight
    ws['J1'].style = highlight
    ws['K1'].style = highlight
    ws['L1'].style = highlight
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['J'].width = 20
    i = 2
    players_num = 0
    for team in teams:
        team_id = team.get('competitor').get('id')
        team_name = team.get('competitor').get('name')
        team_primaryColor = "#" + team.get('competitor').get('primaryColor')
        team_secondaryColor = "#" + team.get('competitor').get('secondaryColor')
        players = team.get('competitor').get('players')
        players_num += len(players)
        for player in players:
            player_id = player.get('player').get('id')
            player_name = player.get('player').get('name')
            homeLocation = player.get('player').get('homeLocation')
            player_number = player.get('player').get('attributes').get('player_number')
            player_role = player.get('player').get('attributes').get('role')
            name = player.get('player').get('givenName') + ' ' + player.get('player').get('familyName')
            nationality = player.get('player').get('nationality')
            ws['A{0}'.format(i)] = i-1
            ws['B{0}'.format(i)] = player_id
            ws['C{0}'.format(i)] = player_name
            ws['D{0}'.format(i)] = homeLocation
            ws['E{0}'.format(i)] = player_number
            ws['F{0}'.format(i)] = player_role
            ws['G{0}'.format(i)] = name
            ws['H{0}'.format(i)] = nationality
            ws['I{0}'.format(i)] = team_id
            ws['J{0}'.format(i)] = team_name
            ws['K{0}'.format(i)] = team_primaryColor
            ws['L{0}'.format(i)] = team_secondaryColor

            ws['A{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['B{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['C{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['D{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['E{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['F{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['G{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['H{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['I{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['J{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['K{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws['L{0}'.format(i)].alignment = Alignment(horizontal='center', vertical='center')
            i += 1
    print('一共有%d个队员' % players_num)
    wb.save("players.xlsx")


if __name__ == '__main__':
    save_excel()

