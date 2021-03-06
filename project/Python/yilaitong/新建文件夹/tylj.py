import openpyxl
import os
from openpyxl.styles import Font
from tkinter import *
import tkinter.filedialog
import warnings

def geshi(ws,name,i):
    wb.create_sheet(index = i+1,title = name)
    wb[name].guess_types = True
    wb[name]['A1'] = '订单编号'
    wb[name]['B1'] = '下单时间'
    wb[name]['C1'] = '商品名称'
    wb[name]['D1'] = '商品数量'
    wb[name]['E1'] = '收货人'
    wb[name]['F1'] = '收货电话'
    wb[name]['G1'] = '收货地址'
    wb[name]['A1'].font = Font(bold=True, size=15)
    wb[name]['B1'].font = Font(bold=True, size=15)
    wb[name]['C1'].font = Font(bold=True, size=15)
    wb[name]['D1'].font = Font(bold=True, size=15)
    wb[name]['E1'].font = Font(bold=True, size=15)
    wb[name]['F1'].font = Font(bold=True, size=15)
    wb[name]['G1'].font = Font(bold=True, size=15)
    
    wb[name].column_dimensions['A'].width = 18
    wb[name].column_dimensions['B'].width = 20
    wb[name].column_dimensions['C'].width = 40
    wb[name].column_dimensions['D'].width = 5
    wb[name].column_dimensions['F'].width = 15
    wb[name].column_dimensions['G'].width = 50
    j = 2
    for a in range(rows-2):
        name1 = '{}'.format(ws['C{}'.format(a+2)].value)
        if names[i] == name1:
            num = ws['E{}'.format(a+2)].value
            if int(num) < 2:
                wb[name]['C{}'.format(j)] = ws['C{}'.format(a+2)].value
                wb[name]['D{}'.format(j)] = ws['E{}'.format(a+2)].value
                name2 = '{}'.format(ws['K{}'.format(a+2)].value)
                if name2 == 'None':
                    for k in range(1,30):
                        name3 = '{}'.format(ws['K{}'.format(a+2-k)].value)
                        if name3 == 'None':
                            continue
                        else:
                            wb[name]['A{}'.format(j)] = ws['A{}'.format(a+2-k)].value
                            wb[name]['B{}'.format(j)] = ws['B{}'.format(a+2-k)].value
                            wb[name]['E{}'.format(j)] = ws['K{}'.format(a+2-k)].value
                            wb[name]['F{}'.format(j)] = ws['L{}'.format(a+2-k)].value
                            wb[name]['G{}'.format(j)] = ws['M{}'.format(a+2-k)].value
                            break
                else:
                    wb[name]['A{}'.format(j)] = ws['A{}'.format(a+2)].value
                    wb[name]['B{}'.format(j)] = ws['B{}'.format(a+2)].value
                    wb[name]['E{}'.format(j)] = ws['K{}'.format(a+2)].value
                    wb[name]['F{}'.format(j)] = ws['L{}'.format(a+2)].value
                    wb[name]['G{}'.format(j)] = ws['M{}'.format(a+2)].value
                j+=1
            else:
                for b in range(int(num)):
                    wb[name]['C{}'.format(j+b)] = ws['C{}'.format(a+2)].value
                    wb[name]['D{}'.format(j+b)] = '1'
                    name2 = '{}'.format(ws['K{}'.format(a+2)].value)
                    if name2 == 'None':
                        for k in range(1,30):
                            name3 = '{}'.format(ws['K{}'.format(a+2-k)].value)
                            if name3 == 'None':
                                continue
                            else:
                                wb[name]['A{}'.format(j+b)] = ws['A{}'.format(a+2-k)].value
                                wb[name]['B{}'.format(j+b)] = ws['B{}'.format(a+2-k)].value
                                wb[name]['E{}'.format(j+b)] = ws['K{}'.format(a+2-k)].value
                                wb[name]['F{}'.format(j+b)] = ws['L{}'.format(a+2-k)].value
                                wb[name]['G{}'.format(j+b)] = ws['M{}'.format(a+2-k)].value
                                break
                    else:
                        wb[name]['A{}'.format(j+b)] = ws['A{}'.format(a+2)].value
                        wb[name]['B{}'.format(j+b)] = ws['B{}'.format(a+2)].value
                        wb[name]['E{}'.format(j+b)] = ws['K{}'.format(a+2)].value
                        wb[name]['F{}'.format(j+b)] = ws['L{}'.format(a+2)].value
                        wb[name]['G{}'.format(j+b)] = ws['M{}'.format(a+2)].value
                j += int(num)

            
        else:
            continue
    wb.save('{}'.format(fname))

if __name__ == '__main__':
    # warnings谨慎使用
    warnings.filterwarnings("ignore")
    root = tkinter.Tk()
    default_dir = r"{}".format(os.getcwd())
    fname = tkinter.filedialog.askopenfilename(title=u"选择文件",initialdir=(os.path.expanduser(default_dir)))
    wb = openpyxl.load_workbook('{}'.format(fname))
    root.withdraw()
    ws = wb['订单列表']
    rows = ws.max_row
    names = []
    for each in ws['C2:C{}'.format(rows-1)]:
        for each_cell in each:
            if each_cell.value not in names:
                names.append(each_cell.value)
    for i in range(len(names)):
        if len(names[i]) < 30:
            name = names[i]
            if len(wb.sheetnames) <= len(names)+1:
                geshi(ws,name,i)
            else:
                print('已录入')
            
        else:
            name = names[i][:20]+' '+names[i][30:]
            if len(wb.sheetnames) <= len(names)+1:
                geshi(ws,name,i)
            else:
                print('已录入')


'''
a = [{'供应商':'张三','商品':['黑贝南瓜 39.9元5斤包邮','红蜜薯5斤39元包邮 48小时内发货']},
     {'供应商':'张三','商品':['黑贝南瓜 39.9元5斤包邮','红蜜薯5斤39元包邮 48小时内发货']}]

'''
            


