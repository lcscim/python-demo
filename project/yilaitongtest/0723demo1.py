import requests
import openpyxl
import yltbj
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
import datetime

def find_url(url):
    headers = {'User-Agent': 'Dalvik/2.1.0 (Linux; U; Android 5.1; MI PAD 2 MIUI/V9.6.1.0.LACCNFD)',
               'Host': 'www.yilaitong.net',
               'Connection':'Keep-Alive',
               'Accept-Encoding': 'gzip'
               }
    req = requests.get(url=url, headers=headers).json()
    name = []
    price = []
    for i in req:
        phonename = i.get('GTitle')
        name.append(phonename)
        price.append(i.get('Price'))

    time = req[0].get('UpTime')
    result = []
    length = len(name)
    for b in range(length):
        result.append([name[b],'','',price[b]])

    return [result,time]
def save_price(a):
    wb = openpyxl.Workbook()
    wb.guess_types = True
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=15)
    highlight.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(highlight)
    
    for i in range(len(a)):
        salename = a[i].get('sale')
        url = a[i].get('url')
        wb.create_sheet(index = i, title = salename)
        wb[salename]['A1'] = "{}型号".format(salename)
        wb[salename]['B1'] = "价格"
        wb[salename]['A1'].style = highlight
        wb[salename]['B1'].style = highlight
        wb[salename].column_dimensions['A'].width = 50
        c = 2
        
        for each in find_url(url)[0]:
            if each[0].find('耳机') == -1:
                if each[0].find('移动') == -1:
                    if each[0].find('电信') == -1:
                        if each[0].find('手环') == -1:
                            wb[salename].append(each)
                            ws = wb[salename]['A{}'.format(c)]
                            ws1 = wb[salename]['D{}'.format(c)]
                            ws.alignment = Alignment(horizontal='center', vertical='center')
                            ws1.alignment = Alignment(horizontal='center', vertical='center')
                            c += 1
                        else:
                            continue
                    else:
                        continue
                else:
                    continue
            else:
                continue
            
        wb[salename]['C1'] = find_url(url)[1]
        wb[salename].merge_cells('C1:D1')
        wb[salename]['C1'].style = highlight
        
    wb.save('{}报价.xlsx'.format(datetime.date.today()))

def get_new_num(sale):
    wb = openpyxl.load_workbook('{}报价.xlsx'.format(datetime.date.today()))
    rows = wb['{}'.format(sale)].max_row
    ws = wb['{}'.format(sale)]['D2:D{}'.format(rows)]
    newnums = []
    if sale == '苹果':
        a = 4000
        b = 100
        c = 150
    elif sale == '华为':
        a = 2000
        b = 100
        c = 150
    elif sale == '小米':
        a = 2000
        b = 100
        c = 150
    elif sale == '魅族':
        a = 1800
        b = 100
        c = 150
    for each in ws:
        for each_cell in each:
            
            if int(each_cell.value) < a:
                if each_cell.value == 0:
                    newnums.append('')
                else:
                    newnums.append(int(each_cell.value) + b)
            else:
                newnums.append(int(each_cell.value) + c)
    
    for j in range(rows-1):
        wb['{}'.format(sale)]['B{}'.format(j+2)] = newnums[j]
        
    wb.save('{}报价.xlsx'.format(datetime.date.today()))

def checkout(sale,url):
    
    wb = openpyxl.load_workbook('{}报价.xlsx'.format(datetime.date.today()))
    wb.guess_types = True
    rows = wb['{}'.format(sale)].max_row
    wsa = wb['{}'.format(sale)]['A2:A{}'.format(rows)]
    wsd = wb['{}'.format(sale)]['D2:D{}'.format(rows)]

    name = []
    zhi = []
    for each in wsa:
        for each_cell in each:
            name.append(each_cell.value)
    for each1 in wsd:
        for each_cell1 in each1:
            zhi.append(each_cell1.value)
    c = rows + 1
    
    for each in find_url(url)[0]:
        if each[0] in name:
            num = name.index(each[0])
            if zhi[num] != 0: 
                wb['{}'.format(sale)]['D{}'.format(num+2)] = (int(each[3])+int(zhi[num]))/2
            else:
                wb['{}'.format(sale)]['D{}'.format(num+2)] = each[3]
        else:
            
            if each[0].find('耳机') == -1:
                if each[0].find('移动') == -1:
                    if each[0].find('电信') == -1:
                        if each[0].find('手环') == -1:
                            wb[sale].append(each)
                            ws = wb[sale]['A{}'.format(c)]
                            ws1 = wb[sale]['D{}'.format(c)]
                            ws.alignment = Alignment(horizontal='center', vertical='center')
                            ws1.alignment = Alignment(horizontal='center', vertical='center')
                            c += 1
                        else:
                            continue
                    else:
                        continue
                else:
                    continue
            else:
                continue


    wb.save('{}报价.xlsx'.format(datetime.date.today()))

def zong_price():
    pinpai = ['苹果','华为','小米','魅族']
    wb = openpyxl.load_workbook('{}报价.xlsx'.format(datetime.date.today()))
    wb.guess_types = True
    wb.create_sheet(index = len(pinpai), title = '总报价')
    for i in range(len(pinpai)):
        a = []
        b = []
        rows = wb['{}'.format(pinpai[i])].max_row
        wsa = wb[pinpai[i]]['A1:A{}'.format(rows)]
        wsb = wb[pinpai[i]]['B1:B{}'.format(rows)]
        for j in wsa:
            for each in j:
                a.append(each.value)
        for k in wsb:
            for each1 in k:
                b.append(each1.value)
        if i == 0:
            for x in range(rows):
                
                wb['总报价']['A{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价']['B{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价'].column_dimensions['A'].width = 40
                wb['总报价']['A{}'.format(x+1)] = a[x]
                wb['总报价']['B{}'.format(x+1)] = b[x]
        if i == 1:
            for x in range(rows):
                
                wb['总报价']['C{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价']['D{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价'].column_dimensions['C'].width = 40
                wb['总报价']['C{}'.format(x+1)] = a[x]
                wb['总报价']['D{}'.format(x+1)] = b[x]
        if i == 2:
            for x in range(rows):
                
                wb['总报价']['E{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价']['F{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价'].column_dimensions['E'].width = 40
                wb['总报价']['E{}'.format(x+1)] = a[x]
                wb['总报价']['F{}'.format(x+1)] = b[x]
        if i == 3:
            for x in range(rows):
                
                wb['总报价']['G{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价']['H{}'.format(x+1)].alignment = Alignment(horizontal='center', vertical='center')
                wb['总报价'].column_dimensions['G'].width = 40
                wb['总报价']['G{}'.format(x+1)] = a[x]
                wb['总报价']['H{}'.format(x+1)] = b[x]
                    
    wb.save('{}报价.xlsx'.format(datetime.date.today()))
           
if __name__ == '__main__':
    
    save_price(yltbj.bjurls())

    for j in range(len(yltbj.jybj())):
        jy_sale = yltbj.jybj()[j].get('sale')
        jy_url = yltbj.jybj()[j].get('url')
        checkout(jy_sale,jy_url)
        get_new_num(jy_sale)
    
    for i in range(len(yltbj.bjurls())):
        sale = yltbj.bjurls()[i].get('sale')
        get_new_num(sale)
    
    zong_price()
    


    
        
